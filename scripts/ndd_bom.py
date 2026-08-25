#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""PCBA BOM (.xlsx) 解析器（通用）。

核心用途：BOM 的 refdes 欄（常見欄名 `Part Reference`）展開後即得
`refdes -> Value / Manufacturer_PN` 的查表。

⚠️ **netlist 有、但該欄沒有的 refdes 就是未貼件 (DNI)**——這是判斷 DNI 的唯一
   可靠方法，不要用「BOM 數量比 netlist 少」去猜是哪幾顆。

⚠️ **標題列位置各檔不同**（有的在第 1 列、有的在第 2 列，欄位順序也不同），
   所以這裡用「哪一列含有 refdes 欄名」自動偵測，**絕不寫死列號或欄號**。

⚠️ **BOM 類型會影響 DNI 判讀**：若拿到的是 SMT BOM，連接器、測試點、鎖孔等
   非 SMT 件本來就不在裡面，不是未貼件。用 `kind` 欄位標註，稽核時分開報。
"""
import io
import os
import re

try:
    import openpyxl
except ImportError:                                  # pragma: no cover
    openpyxl = None

# 常見的 refdes 欄名（大小寫不拘）。找不到時可用 --ref-col 指定。
REF_COL_CANDIDATES = [
    "part reference", "reference", "references", "designator", "designators",
    "refdes", "ref des", "part references", "location",
]
PN_COL_CANDIDATES = ["manufacturer_pn", "manufacturer pn", "mfg pn", "mpn",
                     "manufacturer part number", "part number", "name"]
VAL_COL_CANDIDATES = ["value", "comment", "description"]

_SPLIT_RX = re.compile(r"[,\s;]+")


class Bom(object):
    def __init__(self, path, sheet=None, ref_col=None, kind=""):
        if openpyxl is None:
            raise RuntimeError("需要 openpyxl：pip install openpyxl")
        self.path = path
        self.name = os.path.basename(path)
        self.kind = kind or "未標註"
        self.rows = []
        self.ref = {}
        self.header_row = None
        self.header = []
        self.ref_col_name = None
        self._parse(sheet, ref_col)

    def _parse(self, sheet, ref_col):
        wb = openpyxl.load_workbook(self.path, data_only=True, read_only=True)
        ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
        wanted = [ref_col.lower()] if ref_col else REF_COL_CANDIDATES
        header = None
        idx = None
        for r, row in enumerate(ws.iter_rows(values_only=True), start=1):
            vals = ["" if v is None else str(v).strip() for v in row]
            if header is None:
                low = [v.lower() for v in vals]
                hit = next((w for w in wanted if w in low), None)
                if hit:
                    header, idx = vals, low.index(hit)
                    self.header_row, self.ref_col_name = r, vals[idx]
                continue
            if not any(vals):
                continue
            d = {h: v for h, v in zip(header, vals) if h}
            d["_row"] = r
            self.rows.append(d)
            for rd in self.expand(vals[idx] if idx < len(vals) else ""):
                self.ref[rd] = d
        wb.close()
        if header is None:
            raise ValueError(
                "在 %s 找不到 refdes 欄（試過 %s）。用 --ref-col 指定欄名。"
                % (self.name, ", ".join(wanted)))
        self.header = header

    @staticmethod
    def expand(cell):
        if not cell:
            return []
        return [t for t in _SPLIT_RX.split(cell.strip()) if t]

    def _pick(self, d, cands):
        low = {k.lower(): k for k in d if isinstance(k, str)}
        for c in cands:
            if c in low and d[low[c]]:
                return d[low[c]]
        return ""

    def of(self, refdes):
        return self.ref.get(refdes)

    def pn(self, refdes):
        d = self.ref.get(refdes)
        return self._pick(d, PN_COL_CANDIDATES) if d else None

    def value(self, refdes):
        d = self.ref.get(refdes)
        return self._pick(d, VAL_COL_CANDIDATES) if d else None

    def count_pn(self, pn):
        pn = pn.upper()
        return sum(1 for r in self.ref if (self.pn(r) or "").upper() == pn)

    def refdes_of_pn(self, pn):
        pn = pn.upper()
        return sorted(r for r in self.ref if (self.pn(r) or "").upper() == pn)

    def __repr__(self):
        return "<Bom %s [%s]: %d items / %d refdes (標題列 %d, refdes 欄 '%s')>" % (
            self.name, self.kind, len(self.rows), len(self.ref),
            self.header_row, self.ref_col_name)
